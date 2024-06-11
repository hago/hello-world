#!/usr/bin/env python3
# vim: set fileencoding: UTF-8 -*-

import logging
import os.path
import sys

import openpyxl
import openpyxl.worksheet.worksheet
import openpyxl.worksheet
import openpyxl.workbook
import openpyxl.styles.fonts

from argparse import ArgumentParser
from http.cookiejar import CookieJar
from json import loads
from time import strftime
from urllib.request import quote, Request, build_opener, HTTPCookieProcessor

class productinfo:
    def __init__(self) -> None:
        self.id = None
        self.name = None
        self.price = 0.0
        self.date = None
        self.unit = ''
        self.raw = None

    def __str__(self) -> str:
        return 'id=%s, name=%s, price=%0.2f, date=%s, unit=%s' % (self.id, self.name, self.price, self.date, self.unit)

class querier():
    def __init__(self) -> None:
        self.__URL = '''https://www.cdprice.cn/price/homePageAction!getAllType.do'''
        self.__REFER = '''https://www.cdprice.cn/price/homePageAction!getArticle.do?title=%s'''
        self.__cookiejar = CookieJar()
        self.__opener = build_opener(HTTPCookieProcessor(self.__cookiejar))
        self.__pageloaded = False

    def __loadpage(self, product: str) -> str:
        referurl = self.__REFER % quote(quote(product))
        if not self.__pageloaded:
            header = {
                'Referer': 'https://www.cdprice.cn/',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:127.0) Gecko/20100101 Firefox/127.0'
            }
            req = Request(referurl, method='GET', headers=header)
            with self.__opener.open(req) as rsp:
                pass
            self.__pageloaded = True
        return referurl

    def query(self, product: str)->str:
        data = ('title=%s' % quote(product)).encode('utf-8')
        referurl = self.__loadpage(product)
        header = {
            'Referer': referurl,
            'Host': 'www.cdprice.cn',
            'Origin': 'https://www.cdprice.cn',
            'Content-Type': "application/x-www-form-urlencoded; charset=utf-8",
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:127.0) Gecko/20100101 Firefox/127.0',
            'Content-Length': len(data),
            'X-Requested-With': 'XMLHttpRequest',
            'Accept': 'application/json, text/javascript, */*; q=0.01',
            'Accept-Encoding': 'gzip, deflate, br, zstd',
            'Accept-Language': 'en,zh-CN;q=0.5',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-origin',
            'Cache-Control': 'no-cache',
            'Connection': 'keep-alive'
        }
        req = Request(self.__URL, data, method='POST', headers=header)
        #print(data)
        with self.__opener.open(req) as f:
            data = f.read().decode('UTF-8')
        return data

def parse(jsonstr: str)->list:
    jobj = loads(jsonstr)
    if jobj == None or type(jobj) != dict:
        raise TypeError('Web API respponse malformed')
    if 'list1' not in jobj:
        raise TypeError('"list1" not found')
    list1 = jobj['list1']
    if type(list1) != list:
        raise TypeError('"list1" is NOT a list')
    return [__createproduct(raw) for raw in list1]

def __createproduct(raw: dict)->productinfo:
    logging.debug("parsing: %s", raw)
    p = productinfo()
    p.raw = raw
    p.id = raw['cltpro01id']
    p.name = raw['cltpro01001']
    p.unit = raw['cltpro01005']
    p.date = raw['cltprc05001']
    p.price = raw['cltprc05007']
    return p

class prod2query:
    def __init__(self) -> None:
        self.__data = {}

    def addcategory(self, cat: str):
        self.__data[cat] = []

    def addproduct(self, cat: str, product: str):
        if not cat in self.__data:
            self.addcategory(cat)
        self.__data[cat].append(product)

    def categories(self)->list:
        return list(self.__data.keys())
    
    def products(self, cat: str):
        return self.__data[cat]
    
    '''
    return tuple (品类，名称)
    '''
    def listall(self) -> list:
        l = []
        for cat in self.categories():
            products = self.products(cat)
            for p in products:
                l.append((cat, p))
        return l
    
    def __str__(self) -> str:
        return self.__data.__str__()

def load(infile: str)->list:
    try:
        book = openpyxl.open(infile)
        sh = book[book.sheetnames[0]]
        p = prod2query()
        colnumbers = list(range(2, sh.max_column + 1))
        rownumbers = list(range(2, sh.max_row + 1))
        for i in colnumbers:
            cat = sh.cell(1, i).value
            if cat == None or cat.strip() == '':
                logging.debug("skip column %d", i)
                continue
            logging.debug('col %d: "%s"', i, cat)
            p.addcategory(cat)
            for j in rownumbers:
                name = sh.cell(j, i).value
                if name == None or name.strip() == '':
                    logging.debug("skip column %d row %d", i, j)
                    continue
                p.addproduct(cat, name)
        return p.listall()
    finally:
        book.close

def output(data: list, outfile: str):
    try:
        book = openpyxl.workbook.workbook.Workbook()
        sheetname = '查询结果'
        st: openpyxl.worksheet.worksheet.Worksheet = book.create_sheet(sheetname, 0)
        st.append(['品类', '查询词', 'id', '名称', '价格', '单位', '查询日期'])
        ftbold = openpyxl.styles.fonts.Font(bold=True)
        for row in st["A1": "G1"]:
            for cell in row:
                cell.font = ftbold
        ftred = openpyxl.styles.fonts.Font(color='00ff0000')
        for row in st["A1": "B1"]:
            for cell in row:
                cell.font = ftred
        ftgreen = openpyxl.styles.fonts.Font(color='0000ff00')
        for row in st["C1": "G1"]:
            for cell in row:
                cell.font = ftgreen
        for d in data:
            p: productinfo = d[2]
            st.append([d[0], d[1], p.id, p.name, p.price, p.unit, p.date])
        book.save(outfile)
    finally:
        book.close

if __name__=='__main__':
    parser = ArgumentParser()
    parser.add_argument("-i", "--input", required=True, help="product names to query")
    parser.add_argument("-o", "--output", help="name of excel file containing product info")
    parser.add_argument('-l', '--log-level', default = logging.INFO, help = '''setting log level: CRITICAL, FATAL, ERROR, WARNING, WARN = WARNING, INFO, DEBUG, NOTSET''')
    arg = parser.parse_args()
    logging.basicConfig(level=arg.log_level)
    logging.debug("arg %s", arg)
    f = os.path.realpath(arg.input)
    if not os.path.exists(f):
        logging.error("input file: %s not found", f)
        sys.exit(-1)
    l = load(f)
    logging.debug('parsing result: %s', l)
    q = querier()
    r = []
    for (cate, keyword) in l:
        logging.info('query %s %s', cate, keyword)
        json = q.query(keyword)
        for product in parse(json):
            r.append((cate, keyword, product))
    of = os.path.realpath(strftime('%Y%m%d%H%M%S.xlsx') if arg.output == None else arg.output)
    output(r, of)